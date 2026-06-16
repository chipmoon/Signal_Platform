"""
Nightly Stock Scanner — Composite 6-Month Opportunity Ranker
=============================================================
Quét toàn bộ cổ phiếu TW + VN, tính điểm Composite 6-tháng,
xuất Excel ranking để tìm cổ phiếu tiềm năng tăng giá.

Cách dùng:
    python scripts/nightly_scanner.py              # Scan tất cả
    python scripts/nightly_scanner.py --market VN  # Chỉ VN
    python scripts/nightly_scanner.py --market TW  # Chỉ TW
    python scripts/nightly_scanner.py --top 30     # Top 30 cổ phiếu
    python scripts/nightly_scanner.py --out results/scan_today.xlsx

Composite Score (6-tháng):
    35% Fundamental Quality   (ROE, D/E, EPS growth, P/E, FCF)
    30% Wyckoff + SMC         (phase, structure, signal)
    20% MTF Confluence        (Weekly/Daily/4H alignment)
    15% Elliott Wave Target   (upside potential vs. current price)
"""

from __future__ import annotations

import argparse
import sys
import time
import warnings
from datetime import datetime
from pathlib import Path
from typing import Optional

# ── Project root ──────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
from loguru import logger

# ── Optional rich progress bar ────────────────────────────────────────────────
try:
    from rich.console import Console
    from rich.progress import BarColumn, Progress, SpinnerColumn, TextColumn, TimeElapsedColumn
    from rich.table import Table
    RICH_AVAILABLE = True
    console = Console()
except ImportError:
    RICH_AVAILABLE = False

# ── Analytics engines ─────────────────────────────────────────────────────────
from src.analytics.fundamental_score import compute_fundamental_score
from src.analytics.mtf_confluence import compute_mtf_confluence
from src.strategies.wyckoff_analyzer import WyckoffAnalyzer, WyckoffConfig
from src.strategies.smc_analyzer import SmcAnalyzer, SmcConfig
from src.cache_manager import cache as cm

try:
    from src.strategies.elliott_wave import ElliottWaveAnalyzer
    EW_AVAILABLE = True
except ImportError:
    EW_AVAILABLE = False
    logger.warning("Elliott Wave module not available — EW score will be 0")

try:
    from src.strategies.real_flow_analyzer import RealFlowAnalyzer
    FLOW_AVAILABLE = True
except ImportError:
    FLOW_AVAILABLE = False
    logger.warning("RealFlowAnalyzer not available — flow score will be neutral")

try:
    from src.analytics.position_sizer import compute_position_size
    SIZER_AVAILABLE = True
except ImportError:
    SIZER_AVAILABLE = False
    logger.warning("PositionSizer not available — position sizing will be skipped")

# ── Stock universes ───────────────────────────────────────────────────────────
from src.plugins import registry
from src.plugins.taiwan import TAIWAN_STOCKS  # curated fallback: symbol -> {Name, Sector}

VN_SYMBOLS = [
    # Banks
    "VCB", "BID", "CTG", "TCB", "MBB", "ACB", "VPB", "STB", "HDB",
    "LPB", "SSB", "EIB", "OCB", "TPB", "SHB", "EVF",
    # Real Estate
    "VIC", "VHM", "VRE", "KDH", "NVL", "PDR", "DXG", "IJC", "TDC",
    "SIP", "HDC", "HAG", "LCG", "VPI", "CII", "HDG", "PC1", "VCG", "NLG", "SZC",
    # Oil & Gas
    "PVS", "PVD", "PVT", "PVB", "PVI", "GAS",
    # Industry & Materials
    "HPG", "PLX", "GVR", "BSR", "DGC", "PHR", "DPR", "TRC",
    "BMP", "AAA", "LSS", "PPC", "NT2", "POW", "GEG", "BWE", "KHP",
    "REE", "GEX", "HHV", "NKG", "TLH", "SMC", "HSG",
    # Consumer & Retail
    "SAB", "MSN", "VNM", "MWG", "PNJ", "TLG", "DHC", "DBC", "PAN",
    "VHC", "HAX", "HAH", "VTO", "ASM", "CSV", "BHN",
    # Technology
    "FPT", "CMG", "VNE", "SGT",
    # Aviation & Transport
    "VJC", "GMD", "HVN",
    # Financials / Securities
    "SSI", "VCI", "HCM", "VND", "BSI", "ORS", "CTS", "FTS", "VDS", "MBS",
    # Construction
    "CTD", "FCN", "HBC",
    # Agriculture & Fertilizer
    "AGR", "DPM", "DCM", "DDV",
    # Pharma
    "DHG", "IMP", "TRA",
    # Others
    "BVH", "BCM", "SCS", "TDM", "TV2", "VGC",
]

# ── Composite weights (6-month horizon) ───────────────────────────────────────
# When Foreign Flow data is available: 30%+25%+20%+15%+10% = 100%
# When Flow data unavailable (fail-closed neutral=0.5): effectively same as before
W_FUNDAMENTAL = 0.30
W_WYCKOFF_SMC = 0.25
W_MTF         = 0.20
W_ELLIOTT     = 0.15
W_FLOW        = 0.10   # Foreign Net Flow weight (active only when data available)
# Legacy reference for logging
_WEIGHT_DESC = "30% Fund | 25% W+SMC | 20% MTF | 15% EW | 10% Flow"


# ─────────────────────────────────────────────────────────────────────────────
# Helper: Load price data from cache
# ─────────────────────────────────────────────────────────────────────────────

def _sanitize_price_scale(df: pd.DataFrame, path: Path) -> pd.DataFrame | None:
    """
    Detect and repair scale discontinuities (unit format changes) in price data.
    Returns repaired DataFrame, or None if data is too corrupt to fix.
    """
    if "Close" not in df.columns:
        return df

    close = df["Close"].astype(float)
    log_ret = np.log(close / close.shift(1)).fillna(0)
    big_jumps = log_ret.abs()[log_ret.abs() > 2.0]

    if big_jumps.empty:
        return df  # Clean

    price_cols = [c for c in ["Open", "High", "Low", "Close"] if c in df.columns]
    breaks = []
    for idx in big_jumps.index:
        prev = float(close.iloc[idx - 1])
        if prev == 0:
            continue
        ratio = float(close.iloc[idx]) / prev
        if abs(ratio) > 100 or abs(ratio) < 0.005:
            breaks.append((int(idx), ratio))

    if not breaks:
        return df  # Non-scale jumps (e.g. circuit breakers)

    if len(breaks) > 2:
        # Too corrupt, delete cache so it gets re-fetched
        logger.warning(f"  {path.stem}: corrupt scale ({len(breaks)} breaks), deleting cache")
        path.unlink(missing_ok=True)
        return None

    df_out = df.copy()
    if len(breaks) == 1:
        idx, ratio = breaks[0]
        df_out.loc[:idx - 1, price_cols] = (
            df.loc[:idx - 1, price_cols].astype(float) * ratio
        )
    else:
        idx1, r1 = breaks[0]
        idx2, r2 = breaks[1]
        df_out.loc[idx1:idx2 - 1, price_cols] = (
            df.loc[idx1:idx2 - 1, price_cols].astype(float) * r2
        )
        df_out.loc[:idx1 - 1, price_cols] = (
            df.loc[:idx1 - 1, price_cols].astype(float) * r1 * r2
        )

    # Verify fix worked
    new_log = np.log(df_out["Close"].astype(float) / df_out["Close"].astype(float).shift(1)).fillna(0)
    if new_log.abs().max() > 2.0:
        logger.warning(f"  {path.stem}: scale repair failed, deleting cache")
        path.unlink(missing_ok=True)
        return None

    logger.debug(f"  {path.stem}: auto-repaired {len(breaks)} scale break(s)")
    # Save repaired file back to avoid re-processing on next load
    try:
        df_out.to_parquet(path, engine="pyarrow", index=False)
    except Exception:
        pass
    return df_out


def _load_price_df(symbol: str, market: str) -> Optional[pd.DataFrame]:
    """Load price DataFrame from .cache/prices/ parquet."""
    safe = symbol.replace(".", "_").replace("-", "_")
    path = ROOT / ".cache" / "prices" / f"{safe}_{market}.parquet"
    if not path.exists():
        return None
    try:
        df = pd.read_parquet(path, engine="pyarrow")
        # Normalize columns
        col_map = {c: c.capitalize() for c in df.columns}
        col_map.update({"Adj close": "Close", "Adj Close": "Close"})
        df = df.rename(columns=col_map)
        if "Date" in df.columns:
            df["Date"] = pd.to_datetime(df["Date"])
            df = df.sort_values("Date").reset_index(drop=True)
        # Auto-repair scale discontinuities before returning
        df = _sanitize_price_scale(df, path)
        if df is None:
            return None
        return df if len(df) >= 60 else None
    except Exception as e:
        logger.debug(f"Price load failed for {symbol}: {e}")
        return None


def _fetch_price_from_api(symbol: str, market: str) -> Optional[pd.DataFrame]:
    """Fallback: fetch from yfinance if cache missing."""
    try:
        import yfinance as yf
        if market == "VN":
            ticker_str = f"{symbol}.VN"
        elif market == "TW":
            ticker_str = symbol  # already contains .TW or .TWO
        else:
            ticker_str = symbol

        df = yf.Ticker(ticker_str).history(period="1y", interval="1d", auto_adjust=False)
        if df.empty:
            return None
        df = df.reset_index()
        if "Datetime" in df.columns:
            df = df.rename(columns={"Datetime": "Date"})
        df["Date"] = pd.to_datetime(df["Date"]).dt.tz_localize(None)
        return df[["Date", "Open", "High", "Low", "Close", "Volume"]] if len(df) >= 60 else None
    except Exception:
        return None


def _get_price_df(symbol: str, market: str) -> Optional[pd.DataFrame]:
    """Cache-first price loading."""
    df = _load_price_df(symbol, market)
    if df is not None:
        return df
    logger.debug(f"  {symbol}: no cache, trying API...")
    return _fetch_price_from_api(symbol, market)


# ─────────────────────────────────────────────────────────────────────────────
# Scoring engines
# ─────────────────────────────────────────────────────────────────────────────

def score_fundamental(symbol: str, market: str) -> tuple[float, int, str]:
    """
    Returns (normalized_score 0-1, raw_score 0-100, grade).
    Uses cache-first strategy inside compute_fundamental_score.
    """
    try:
        fs = compute_fundamental_score(symbol, market)
        if fs.data_coverage < 40 or str(fs.grade).upper() == "N/A":
            return 0.5, 0, "N/A"
        return fs.total_score / 100.0, fs.total_score, fs.grade
    except Exception as e:
        logger.debug(f"  {symbol} fundamental error: {e}")
        return 0.5, 0, "N/A"


def score_wyckoff_smc(df: pd.DataFrame) -> tuple[float, str, float]:
    """
    Returns (normalized_score 0-1, wyckoff_phase, smc_score_raw).
    Wyckoff: Accumulation/Markup → bullish, Distribution/Markdown → bearish.
    SMC: smc_score in [-1, +1] range.
    """
    try:
        wa = WyckoffAnalyzer(WyckoffConfig())
        wy = wa.analyze_current_state(df)
        w_score = float(wy.get("score", 0.0))        # typically -1 to +1
        w_phase = str(wy.get("phase", "Unknown"))

        smc = SmcAnalyzer(SmcConfig()).get_current_state(df)
        smc_raw = float(smc.get("smc_score", 0.0))   # typically -1 to +1

        # Combine: 60% Wyckoff + 40% SMC, normalize to 0-1
        combined = (w_score * 0.6 + smc_raw * 0.4 + 1.0) / 2.0  # map [-1,1] → [0,1]
        combined = max(0.0, min(1.0, combined))
        return combined, w_phase, smc_raw
    except Exception as e:
        logger.debug(f"  Wyckoff/SMC error: {e}")
        return 0.5, "Unknown", 0.0  # neutral default


def score_mtf(symbol: str, market: str, df_daily: pd.DataFrame) -> tuple[float, str, int]:
    """
    Returns (normalized_score 0-1, confluence_label, alignment 0-3).
    confluence_score range is -9 to +9 → normalize to 0-1.
    """
    try:
        mtf = compute_mtf_confluence(symbol, market, df_daily=df_daily)
        raw = float(mtf.get("confluence_score", 0.0))  # -9 to +9
        normalized = (raw + 9.0) / 18.0                # → 0 to 1
        normalized = max(0.0, min(1.0, normalized))
        label = mtf.get("confluence_label", "Neutral")
        alignment = int(mtf.get("alignment", 0))
        return normalized, label, alignment
    except Exception as e:
        logger.debug(f"  MTF error: {e}")
        return 0.5, "Unknown", 0


def score_elliott(df: pd.DataFrame) -> tuple[float, float]:
    """
    Returns (normalized_score 0-1, upside_pct).
    Measures how much upside the Elliott Wave target implies vs. current price.
    """
    if not EW_AVAILABLE or df is None or len(df) < 50:
        return 0.5, 0.0  # neutral default

    try:
        ew = ElliottWaveAnalyzer()
        # Try weekly resample first
        df_ew = df.copy()
        if "Date" in df_ew.columns:
            df_ew = df_ew.set_index("Date")
        df_weekly = df_ew.resample("W").agg(
            {"Open": "first", "High": "max", "Low": "min", "Close": "last", "Volume": "sum"}
        ).dropna().reset_index()
        df_weekly.columns = [str(c) for c in df_weekly.columns]

        state = ew.get_current_state(df_weekly) if len(df_weekly) >= 20 else {}
        target = float(state.get("target_price", 0.0))
        curr = float(df["Close"].iloc[-1]) if "Close" in df.columns else 0.0
        bias = str(state.get("bias", "Neutral"))
        confidence = float(state.get("confidence", 0)) / 100.0

        if target > 0 and curr > 0:
            upside_pct = (target - curr) / curr * 100.0
        else:
            upside_pct = 0.0

        # Score: bullish bias + upside potential, weighted by confidence
        if "Bullish" in bias and upside_pct > 0:
            raw = min(upside_pct / 50.0, 1.0) * confidence  # 50% upside = max score
            score = 0.5 + raw * 0.5   # range 0.5-1.0 for bullish
        elif "Bearish" in bias:
            score = max(0.0, 0.5 - confidence * 0.3)
        else:
            score = 0.5

        return score, upside_pct
    except Exception as e:
        logger.debug(f"  Elliott Wave error: {e}")
        return 0.5, 0.0


# ─────────────────────────────────────────────────────────────────────────────
# Main scanner
# ─────────────────────────────────────────────────────────────────────────────

def scan_symbol(symbol: str, market: str, name: str, sector: str) -> Optional[dict]:
    """Run full composite scoring for one symbol. Returns result dict or None."""
    # 1. Load price data
    df = _get_price_df(symbol, market)
    if df is None:
        logger.warning(f"  {symbol}: no price data — skipping")
        return None

    current_price = float(df["Close"].iloc[-1]) if "Close" in df.columns else 0.0
    if current_price <= 0:
        return None

    # 2. Score each component
    fund_norm, fund_raw, fund_grade = score_fundamental(symbol, market)
    wy_smc_norm, wy_phase, smc_raw = score_wyckoff_smc(df)
    mtf_norm, mtf_label, mtf_align = score_mtf(symbol, market, df)
    ew_norm, upside_pct = score_elliott(df)

    # 2.5 Compute Short-term (1-3M) Institutional Recommendation
    short_term_rating = "Neutral (Trung lập)"
    try:
        from src.strategies.alpha_scanner import AlphaScannerEngine
        if market == "VN":
            alpha_symbol = f"{symbol}.VN" if not symbol.endswith(".VN") else symbol
            bench_sym = "VNINDEX"
        elif market == "TW":
            # 8096 trades on TWO board, all others on TW
            alpha_symbol = f"{symbol}.TWO" if symbol == "8096" else f"{symbol}.TW"
            bench_sym = "2330.TW"
        else:
            alpha_symbol = symbol
            bench_sym = "^GSPC"

        alpha_res = AlphaScannerEngine._scan_tier2({"symbol": alpha_symbol, "market": market, "benchmark": bench_sym})
        if alpha_res:
            rec = alpha_res.get("recommendation", "WATCH")
            if rec == "STRONG BUY":
                short_term_rating = "⭐⭐⭐ Strong Buy (Sóng ngắn)"
            elif rec == "BUY":
                short_term_rating = "⭐⭐ Buy (Sóng ngắn)"
            elif rec == "WATCH":
                short_term_rating = "Watch (Đang tích lũy)"
            elif rec == "AVOID":
                short_term_rating = "Avoid (Giảm mạnh)"
            else:
                short_term_rating = "Neutral (Trung lập)"
    except Exception as e:
        logger.debug(f"Alpha scanner scan failed for {symbol}: {e}")

    # 2.6 Foreign Net Flow Score (real_flow_analyzer — fail-closed neutral if no data)
    flow_norm      = 0.5   # neutral default: flow weight contributes equally to all
    flow_available = False
    flow_signal    = "N/A"
    flow_net_3d    = 0.0
    if FLOW_AVAILABLE and market == "VN":  # Foreign flow data only for VN market
        try:
            rfa = RealFlowAnalyzer()
            df_flow = rfa.generate_signals(df)
            if df_flow["real_flow_available"].any():
                flow_available = True
                # Average of last 3 sessions for stability
                raw_3d = float(df_flow["real_flow_score"].iloc[-3:].mean())
                flow_norm = max(0.0, min(1.0, (raw_3d + 1.0) / 2.0))  # map [-1,1]→[0,1]
                flow_net_3d = float(df_flow["foreign_net_flow_ratio"].iloc[-3:].mean())
                if raw_3d > 0.20:
                    flow_signal = "KN Mua rong"
                elif raw_3d < -0.20:
                    flow_signal = "KN Ban rong"
                else:
                    flow_signal = "Trung lap"
        except Exception as e:
            logger.debug(f"  {symbol} flow error: {e}")

    # 3. Composite score (0-100)
    # When flow data available: weighted 5-component score
    # When flow unavailable: flow_norm=0.5 (neutral) — math is equivalent to 4-component
    composite = (
        fund_norm   * W_FUNDAMENTAL +
        wy_smc_norm * W_WYCKOFF_SMC +
        mtf_norm    * W_MTF +
        ew_norm     * W_ELLIOTT +
        flow_norm   * W_FLOW
    ) * 100.0

    # 4. Opportunity rating
    if composite >= 72:
        rating = "⭐⭐⭐ Strong Buy"
    elif composite >= 60:
        rating = "⭐⭐ Watch"
    elif composite >= 45:
        rating = "⭐ Neutral"
    else:
        rating = "⚠️ Avoid"

    # 3.5 Position Sizing
    pos = {
        "recommended_allocation_pct": 0.0, "recommended_vnd_million": 0.0,
        "recommended_shares": 0, "stop_loss_pct": 5.0,
        "risk_level": "N/A", "annualized_vol_pct": 0.0,
    }
    if SIZER_AVAILABLE:
        try:
            pos = compute_position_size(
                symbol=symbol,
                current_price=current_price,
                mid_rating=rating,
                short_rating=short_term_rating,
                df_price=df,
            )
        except Exception as e:
            logger.debug(f"  {symbol} position sizing error: {e}")

    return {
        "Symbol":           symbol,
        "Name":             name,
        "Market":           market,
        "Sector":           sector,
        "Price":            round(current_price, 2),
        # Composite
        "Composite Score":  round(composite, 1),
        "Mid-term Rating (6M)": rating,
        "Short-term Rating (1-3M)": short_term_rating,
        # Fundamental
        "Fund Score":       fund_raw,
        "Fund Grade":       fund_grade,
        # Wyckoff + SMC
        "Wyckoff Phase":    wy_phase,
        "SMC Score":        round(smc_raw, 3),
        "W+SMC Score":      round(wy_smc_norm * 100, 1),
        # MTF
        "MTF Label":        mtf_label.replace("\U0001f7e2 ", "").replace("\U0001f534 ", "").replace("\u26aa ", ""),
        "MTF Alignment":    mtf_align,
        "MTF Score":        round(mtf_norm * 100, 1),
        # Elliott
        "EW Upside %":      round(upside_pct, 1),
        "EW Score":         round(ew_norm * 100, 1),
        # Foreign Net Flow (new)
        "Flow Signal":            flow_signal,
        "KN Net Flow (3d)": round(flow_net_3d, 4),
        "Flow Score":             round(flow_norm * 100, 1),
        "Flow Data":              "Yes" if flow_available else "No",
        # Position Sizing (new)
        "Alloc % (1B VND)":     pos["recommended_allocation_pct"],
        "Amount (M VND)":         pos["recommended_vnd_million"],
        "Shares (lot-100)":       pos["recommended_shares"],
        "Stop Loss %":            pos["stop_loss_pct"],
        "Risk Level":             pos["risk_level"],
        "Volatility (Ann.)":      pos["annualized_vol_pct"],
    }


def _build_universe(market_filter: Optional[str]) -> list[tuple[str, str, str, str]]:
    """Returns list of (symbol, market, name, sector)."""
    universe = []

    if market_filter in (None, "TW"):
        tw_assets = []
        try:
            provider = registry.get("TW")
            if provider:
                tw_assets = provider.search_assets("", limit=5000)
        except Exception as exc:
            logger.debug(f"TW universe provider load failed: {exc}")

        if tw_assets:
            for asset in tw_assets:
                universe.append((asset.symbol, "TW", asset.name, asset.sector or "Other"))
        else:
            for sym, info in TAIWAN_STOCKS.items():
                universe.append((sym, "TW", info.get("Name", sym), info.get("Sector", "Other")))

    if market_filter in (None, "VN"):
        for sym in VN_SYMBOLS:
            universe.append((sym, "VN", sym, "VN Stock"))

    return universe


def _style_excel(writer: pd.ExcelWriter, df_all: pd.DataFrame, df_vn: pd.DataFrame, df_tw: pd.DataFrame):
    """Apply conditional formatting and column widths to Excel sheets."""
    try:
        from openpyxl.formatting.rule import ColorScaleRule
        from openpyxl.styles import Alignment, Font, PatternFill

        COLOR_SCALE = ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B",
        )
        # Header fill for all sheets
        HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        HEADER_FONT = Font(bold=True, color="FFFFFF")

        all_sheets = [("All Stocks", df_all), ("VN Ranking", df_vn), ("TW Ranking", df_tw)]
        if "⚡ Short-term Picks" in writer.sheets:
            # Pull df for short-term picks sheet (already written; reconstruct from df_all)
            pass  # styling handled in same loop via sheet name lookup

        for sheet_name, df in all_sheets:
            if df is None or df.empty or sheet_name not in writer.sheets:
                continue
            ws = writer.sheets[sheet_name]
            cols = list(df.columns)

            # Style header row
            for cell in ws[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            # Auto-width columns
            for col_idx, col in enumerate(cols, 1):
                try:
                    max_len = max(len(str(col)), df[col].astype(str).str.len().max())
                except Exception:
                    max_len = len(str(col))
                ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_len + 2, 32)

            # Color-scale: Composite Score
            for score_col in ("Composite Score", "Fund Score", "MTF Score", "EW Score"):
                if score_col in cols:
                    idx = cols.index(score_col) + 1
                    ltr = ws.cell(1, idx).column_letter
                    ws.conditional_formatting.add(
                        f"{ltr}2:{ltr}{len(df)+1}",
                        ColorScaleRule(
                            start_type="min", start_color="F8696B",
                            mid_type="percentile", mid_value=50, mid_color="FFEB84",
                            end_type="max", end_color="63BE7B",
                        )
                    )

        # Style ⚡ Short-term Picks sheet if present
        if "⚡ Short-term Picks" in writer.sheets:
            ws_st = writer.sheets["⚡ Short-term Picks"]
            for cell in ws_st[1]:
                cell.fill = PatternFill(start_color="0D3349", end_color="0D3349", fill_type="solid")
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            for col_idx in range(1, ws_st.max_column + 1):
                col_letter = ws_st.cell(1, col_idx).column_letter
                ws_st.column_dimensions[col_letter].width = 22

    except Exception as e:
        logger.debug(f"Excel styling error (non-critical): {e}")


def run_scan(market_filter: Optional[str], top_n: Optional[int], output_path: Path) -> pd.DataFrame:
    universe = _build_universe(market_filter)
    total = len(universe)
    results = []
    failed = []

    logger.info(f"🔍 Starting scan: {total} symbols | Market: {market_filter or 'ALL'}")
    logger.info(f"   Weights — {_WEIGHT_DESC}")
    logger.info(f"   Flow module: {'[Active]' if FLOW_AVAILABLE else '[Disabled - fail-closed neutral]'}")
    logger.info(f"   Sizer module: {'[Active]' if SIZER_AVAILABLE else '[Disabled]'}")

    if RICH_AVAILABLE:
        progress_ctx = Progress(
            SpinnerColumn(),
            TextColumn("[bold blue]{task.description}"),
            BarColumn(),
            TextColumn("[progress.percentage]{task.percentage:>3.0f}%"),
            TextColumn("• {task.completed}/{task.total}"),
            TimeElapsedColumn(),
            console=console,
            refresh_per_second=4,
        )
    else:
        progress_ctx = None

    def _process():
        for i, (symbol, market, name, sector) in enumerate(universe, 1):
            desc = f"[{i:3d}/{total}] {symbol:<12} ({market})"
            if RICH_AVAILABLE:
                task = progress_ctx.add_task(desc, total=1)

            try:
                result = scan_symbol(symbol, market, name, sector)
                if result:
                    results.append(result)
                    if RICH_AVAILABLE:
                        progress_ctx.update(task, advance=1, description=f"✅ {desc} → {result['Composite Score']:.0f}")
                else:
                    failed.append(symbol)
                    if RICH_AVAILABLE:
                        progress_ctx.update(task, advance=1, description=f"⚠️  {desc} → no data")
            except Exception as e:
                failed.append(symbol)
                logger.debug(f"  SCAN ERROR {symbol}: {e}")
                if RICH_AVAILABLE:
                    progress_ctx.update(task, advance=1, description=f"❌ {desc}")

            # Small throttle to avoid rate limiting
            time.sleep(0.3)

    if RICH_AVAILABLE:
        with progress_ctx:
            _process()
    else:
        _process()

    if not results:
        logger.error("No results collected — check cache or connectivity")
        return pd.DataFrame()

    # Sort by Composite Score descending
    df = pd.DataFrame(results).sort_values("Composite Score", ascending=False).reset_index(drop=True)
    df.insert(0, "Rank", range(1, len(df) + 1))

    df_vn = df[df["Market"] == "VN"].reset_index(drop=True)
    df_vn.insert(0, "VN Rank", range(1, len(df_vn) + 1))

    df_tw = df[df["Market"] == "TW"].reset_index(drop=True)
    df_tw.insert(0, "TW Rank", range(1, len(df_tw) + 1))

    df_top = df.head(top_n) if top_n else df

    # ── Short-term Picks: sorted by Short-term rating strength ────────────────
    _ST_ORDER = {
        "⭐⭐⭐ Strong Buy (Sóng ngắn)": 0,
        "⭐⭐ Buy (Sóng ngắn)": 1,
        "Watch (Đang tích lũy)": 2,
        "Neutral (Trung lập)": 3,
        "Avoid (Giảm mạnh)": 4,
    }
    df["_st_order"] = df["Short-term Rating (1-3M)"].map(_ST_ORDER).fillna(3)
    df_st = df.sort_values(["_st_order", "Composite Score"], ascending=[True, False]).drop(columns=["_st_order"]).reset_index(drop=True)
    df_st.insert(0, "ST Rank", range(1, len(df_st) + 1))
    # Remove the helper col from main df too
    df = df.drop(columns=["_st_order"], errors="ignore")

    # ── Export to Excel ───────────────────────────────────────────────────────
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Sheet 1: Top Investment Picks (Mid-term 6M, by Composite Score)
        df_top.to_excel(writer, sheet_name="🏆 Top Picks (6M)", index=False)

        # Sheet 2: Short-term Speculative Picks (1-3M, by ST Rating)
        df_st.to_excel(writer, sheet_name="⚡ Short-term Picks", index=False)

        # Sheet 3: VN Ranking
        df_vn.to_excel(writer, sheet_name="VN Ranking", index=False)

        # Sheet 4: TW Ranking
        df_tw.to_excel(writer, sheet_name="TW Ranking", index=False)

        # Sheet 5: Position Sizing Guide (sorted by allocation)
        pos_cols = [
            "Symbol", "Market", "Sector", "Price",
            "Mid-term Rating (6M)", "Short-term Rating (1-3M)",
            "Composite Score", "Volatility (Ann.)", "Risk Level",
            "Alloc % (1B VND)", "Amount (M VND)", "Shares (lot-100)",
            "Stop Loss %",
            "Flow Signal", "KN Net Flow (3d)", "Flow Score", "Flow Data",
        ]
        pos_cols_avail = [c for c in pos_cols if c in df.columns]
        df_pos = df[pos_cols_avail].sort_values(
            "Alloc % (1B VND)", ascending=False, na_position="last"
        ).reset_index(drop=True)
        df_pos.to_excel(writer, sheet_name="Position Sizing", index=False)

        # Sheet 6: All Stocks (full)
        df.to_excel(writer, sheet_name="All Stocks", index=False)

        # Sheet 7: Scan Info / Metadata
        st_strong_cnt = int((df["Short-term Rating (1-3M)"].str.contains("Strong Buy", na=False)).sum())
        mt_strong_cnt = int((df["Mid-term Rating (6M)"].str.contains("Strong Buy", na=False)).sum())
        meta = pd.DataFrame([{
            "Run Date": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "Total Scanned": total,
            "Successful": len(results),
            "Failed": len(failed),
            "Failed Symbols": ", ".join(failed[:30]),
            # Mid-term (6M) weights
            "[6M] Fundamental Weight": f"{W_FUNDAMENTAL*100:.0f}%",
            "[6M] Wyckoff+SMC Weight": f"{W_WYCKOFF_SMC*100:.0f}%",
            "[6M] MTF Weight":         f"{W_MTF*100:.0f}%",
            "[6M] Elliott Wave Weight": f"{W_ELLIOTT*100:.0f}%",
            "[6M] Flow Weight":         f"{W_FLOW*100:.0f}% (active when KN data available)",
            "[6M] Strong Buy Threshold": "Composite Score ≥ 72",
            "[6M] Watch Threshold":      "Composite Score ≥ 60",
            "[6M] Strong Buy Count":     mt_strong_cnt,
            # Short-term (1-3M) engine
            "[1-3M] Engine": "AlphaScannerEngine Tier-2 (AI Predictor + Institutional Flow)",
            "[1-3M] Strong Buy Threshold": "AI Score ≥ 0.45 & Confidence ≥ 0.25",
            "[1-3M] Strong Buy Count": st_strong_cnt,
        }])
        meta.T.reset_index().rename(columns={"index": "Parameter", 0: "Value"}).to_excel(
            writer, sheet_name="Scan Info", index=False
        )

        _style_excel(writer, df, df_vn, df_tw)

    logger.success(f"✅ Scan complete! {len(results)}/{total} symbols scored")
    logger.success(f"📊 Excel saved → {output_path}")

    # ── Print top 15 to terminal ──────────────────────────────────────────────
    top15 = df.head(15)
    if RICH_AVAILABLE:
        table = Table(title=f"🏆 Top 15 — Composite Score [{datetime.now().strftime('%Y-%m-%d')}]",
                      show_header=True, header_style="bold cyan")
        for col in ["Rank", "Symbol", "Market", "Composite Score", "Mid-term Rating (6M)", "Short-term Rating (1-3M)",
                    "Fund Grade", "Wyckoff Phase", "MTF Label", "EW Upside %"]:
            table.add_column(col, justify="right" if col in ("Rank", "Composite Score", "EW Upside %") else "left")
        for _, row in top15.iterrows():
            table.add_row(
                str(int(row["Rank"])), row["Symbol"], row["Market"],
                f"{row['Composite Score']:.1f}", row["Mid-term Rating (6M)"], row["Short-term Rating (1-3M)"],
                row["Fund Grade"], str(row["Wyckoff Phase"]),
                row["MTF Label"], f"{row['EW Upside %']:+.1f}%",
            )
        console.print(table)
    else:
        print(f"\n{'='*80}")
        print(f"TOP 15 — Composite Score [{datetime.now().strftime('%Y-%m-%d')}]")
        print(f"{'='*80}")
        print(top15[["Rank", "Symbol", "Market", "Composite Score", "Mid-term Rating (6M)", "Short-term Rating (1-3M)",
                      "Fund Grade", "Wyckoff Phase", "MTF Label", "EW Upside %"]].to_string(index=False))
        print(f"{'='*80}\n")

    return df


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Nightly Stock Scanner — Composite 6-Month Ranker",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("--market", choices=["VN", "TW"], default=None,
                        help="Filter to specific market (default: both)")
    parser.add_argument("--top", type=int, default=None,
                        help="Show/export only top N stocks in 'Top Picks' sheet")
    parser.add_argument("--out", type=str, default=None,
                        help="Output Excel path (default: output/scan_YYYYMMDD.xlsx)")
    parser.add_argument("--no-excel", action="store_true",
                        help="Skip Excel export, print to terminal only")
    args = parser.parse_args()

    # Default output path
    date_str = datetime.now().strftime("%Y%m%d_%H%M")
    if args.out:
        output_path = Path(args.out)
    else:
        output_path = ROOT / f"scan_{date_str}.xlsx"

    start = time.time()
    df = run_scan(
        market_filter=args.market,
        top_n=args.top,
        output_path=output_path,
    )
    elapsed = time.time() - start

    if not df.empty:
        strong_buys = df[df["Mid-term Rating (6M)"].str.contains("Strong Buy", na=False)]
        watch = df[df["Mid-term Rating (6M)"].str.contains("Watch", na=False)]
        logger.info(f"⏱️  Total time: {elapsed/60:.1f} min")
        logger.info(f"⭐⭐⭐ Mid-term Strong Buy candidates: {len(strong_buys)}")
        logger.info(f"⭐⭐  Watch list: {len(watch)}")

        # Print short-term specs
        st_strong_buys = df[df["Short-term Rating (1-3M)"].str.contains("Strong Buy", na=False)]
        logger.info(f"⚡ Short-term Strong Buy (1-3M Speculation): {len(st_strong_buys)}")

        if not strong_buys.empty:
            syms = strong_buys["Symbol"].tolist()
            logger.success(f"💡 Mid-term Strong Buy: {', '.join(syms[:20])}")
        if not st_strong_buys.empty:
            syms_st = st_strong_buys["Symbol"].tolist()
            logger.success(f"🔥 Short-term Speculative Strong Buy: {', '.join(syms_st[:20])}")
        
        logger.info("→ Copy any symbol above and paste into AI-Forecast to deep-dive analysis")


if __name__ == "__main__":
    main()
